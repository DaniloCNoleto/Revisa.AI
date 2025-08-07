# Scripts/Gerenciador_Revisores_Estruturados.py
# VERSÃO FINAL - Orquestra os três scripts e inclui logging em planilha mestra.

import os
import subprocess
import sys
import smtplib
import time
import shutil
import argparse 
import multiprocessing
from multiprocessing import Pool
from email.message import EmailMessage
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from docx import Document

# --- Bloco de Inicialização Django ---
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.append(str(PROJECT_ROOT))

# --- CARREGAR VARIÁVEIS DE AMBIENTE PRIMEIRO ---
load_dotenv(PROJECT_ROOT / ".env")

import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "dossel_project.settings")
django.setup()
from django.db import close_old_connections
close_old_connections()
from revisor.models import QueueEntry
from django.conf import settings


PASTA_SCRIPTS = "Scripts"
SCRIPT_REVISAO_COMPLETA = os.path.join(PASTA_SCRIPTS, "Processador_Unificado.py")
SCRIPT_REVISAO_SIMPLES = os.path.join(PASTA_SCRIPTS, "Revisor_Textual_Estruturado.py")
SCRIPT_VERIFICADOR_INCONSISTENCIAS = os.path.join(PASTA_SCRIPTS, "verificador_contradicoes_globais.py")

PASTA_ENTRADA = settings.PASTA_ENTRADA
PASTA_SAIDA   = settings.PASTA_SAIDA
PASTA_CANCELAMENTO = settings.PASTA_CANCELAMENTO 

SENHA_APP = os.getenv("SENHA_APP")
EMAIL_REMETENTE = os.getenv("EMAIL_REMETENTE")
EMAIL_MESTRE = "n.danilo@dosselambiental.com.br"
PLANILHA_MESTRA_NOME = PROJECT_ROOT / "relatorio_geral_revisoes.xlsx"

# --- Funções de Controle e Auxiliares ---

def foi_cancelado(rev_id: int) -> bool:
    """Verifica se existe o arquivo cancel_<id>.txt na PASTA_CANCELAMENTO."""
    if not rev_id: return False
    sinal = Path(PASTA_CANCELAMENTO) / f"cancel_{rev_id}.txt"
    if sinal.exists():
        print(f" Cancelamento detectado para a revisão {rev_id}.")
        try:
            sinal.unlink()
        except OSError as e:
            print(f"  Não foi possível remover o arquivo de sinal: {e}")
        return True
    return False

def atualizar_status_progresso(percentual, id_revisao=None):
    """Atualiza o progresso no banco de dados do Django."""
    if not id_revisao: return
    try:
        progresso_int = int(percentual)
        QueueEntry.objects.filter(id=id_revisao).update(progress=progresso_int)
        if progresso_int >= 100:
            QueueEntry.objects.filter(id=id_revisao).update(status='concluido')
        print(f" Progresso atualizado para: {progresso_int}% (ID da Revisão: {id_revisao})")
    except Exception as e:
        print(f" ERRO CRÍTICO ao atualizar progresso no banco de dados para a revisão ID {id_revisao}: {e}")

def eh_documento_valido(nome_arquivo):
    """Verifica se o arquivo é um .docx válido."""
    return nome_arquivo.endswith(".docx") and not nome_arquivo.startswith("~$")

def registrar_documento(nome):
    """Registra um documento como processado com sucesso."""
    with open(PROJECT_ROOT / "documentos_processados.txt", "a", encoding="utf-8") as f:
        f.write(nome + "\n")

def enviar_email_final(assunto, corpo_resumo, email_destino_dinamico):
    """Envia um e-mail de notificação para o usuário final."""
    if not all([EMAIL_REMETENTE, email_destino_dinamico, SENHA_APP]):
        print("  Credenciais de e-mail ou destinatário não configurados. Pulando envio.")
        return
    msg = EmailMessage()
    msg["Subject"] = assunto
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = email_destino_dinamico
    msg.set_content(corpo_resumo)
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_REMETENTE, SENHA_APP)
            smtp.send_message(msg)
        print(f"E-mail de conclusão enviado com sucesso para {email_destino_dinamico}.")
    except Exception as e:
        print(f" Erro ao enviar e-mail: {e}")

def obter_contagem_paginas(caminho_arquivo):
    """Tenta ler a contagem de páginas dos metadados do documento .docx."""
    try:
        doc = Document(caminho_arquivo)
        # A propriedade 'pages' pode não estar presente; fallback para 1.
        return doc.core_properties.pages or 1
    except Exception:
        return 1

def calcular_custo_e_tokens(nome_base_doc, modo):
    """Calcula o custo e os tokens lendo a planilha de resultado apropriada."""
    total_tokens_in, total_tokens_out, total_custo_brl = 0, 0, 0.0
    
    planilhas_possiveis = {
        'completa': Path(PASTA_SAIDA) / nome_base_doc / "avaliacao_completa.xlsx",
        'simples': Path(PASTA_SAIDA) / nome_base_doc / "relatorio_revisao_simples.xlsx",
        'inconsistencias': Path(PASTA_SAIDA) / nome_base_doc / f"relatorio_inconsistencias_{nome_base_doc}.xlsx",
    }
    path_planilha = planilhas_possiveis.get(modo)

    if path_planilha and path_planilha.is_file():
        try:
            wb = load_workbook(path_planilha)
            if "Resumo_Custos" in wb.sheetnames:
                ws = wb["Resumo_Custos"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and isinstance(row[1], (int, float)): total_tokens_in += int(row[1])
                    if row[2] and isinstance(row[2], (int, float)): total_tokens_out += int(row[2])
                    if row[4] and isinstance(row[4], (int, float)): total_custo_brl += row[4]
        except Exception as e:
            print(f"Aviso: Não foi possível ler a planilha de resumo para '{nome_base_doc}'. Erro: {e}")
            
    return total_tokens_in, total_tokens_out, total_custo_brl

def atualizar_planilha_mestra(dados_linha):
    """Cria ou atualiza a planilha mestra com os dados da última execução."""
    print("Atualizando planilha mestra de rastreamento...")
    cabecalho = ["Data e Hora", "Nome do Documento", "Qtd. Páginas", "Tipo de Revisão", "Tempo de Execução (s)", "Custo (BRL)", "Caminho dos Arquivos"]
    
    try:
        if PLANILHA_MESTRA_NOME.exists():
            wb = load_workbook(PLANILHA_MESTRA_NOME)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(cabecalho)
        
        ws.append(dados_linha)
        wb.save(PLANILHA_MESTRA_NOME)
        print("Planilha mestra atualizada com sucesso.")
        return True
    except Exception as e:
        print(f" ERRO ao atualizar a planilha mestra: {e}")
        return False

def enviar_planilha_mestra_por_email():
    """Envia a planilha mestra como anexo para o e-mail fixo."""
    if not all([EMAIL_REMETENTE, EMAIL_MESTRE, SENHA_APP]):
        print("  Credenciais de e-mail ou destinatário mestre não configurados. Pulando envio da planilha.")
        return
    if not PLANILHA_MESTRA_NOME.exists():
        print("  Arquivo da planilha mestra não encontrado para envio.")
        return
        
    print(f"Enviando planilha mestra para {EMAIL_MESTRE}...")
    msg = EmailMessage()
    msg["Subject"] = f"Relatório de Revisões Atualizado - {time.strftime('%Y-%m-%d %H:%M')}"
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = EMAIL_MESTRE
    msg.set_content("Segue em anexo a planilha de rastreamento de revisões atualizada.")
    
    with open(PLANILHA_MESTRA_NOME, 'rb') as f:
        msg.add_attachment(f.read(), maintype='application', subtype='octet-stream', filename=PLANILHA_MESTRA_NOME.name)
    
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_REMETENTE, SENHA_APP)
            smtp.send_message(msg)
        print("Planilha mestra enviada com sucesso.")
    except Exception as e:
        print(f" Erro ao enviar a planilha mestra por e-mail: {e}")

def rodar_script(script_e_argumentos):
    """Executa um script filho como um subprocesso, imprimindo sua saída em tempo real."""
    script, nome_arquivo_com_extensao, id_revisao = script_e_argumentos
    comando = [sys.executable, '-u', script, nome_arquivo_com_extensao, '--id_revisao', str(id_revisao)]
    print(f"\n Executando: {' '.join(comando)}")

    # Usamos Popen para ter controle em tempo real sobre o processo
    proc = subprocess.Popen(
        comando,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="ignore",
    )

    # --- LÓGICA DE LEITURA EM TEMPO REAL ---
    # Lê e imprime a saída padrão (stdout) linha por linha, conforme ela é gerada
    while proc.poll() is None:
        stdout_line = proc.stdout.readline()
        if stdout_line:
            print(stdout_line, end='')
            sys.stdout.flush() # Força a escrita imediata no log

        # Verificação de cancelamento
        if foi_cancelado(id_revisao):
            print(" Cancelamento solicitado; encerrando subprocesso…")
            proc.terminate()
            # Espera um pouco para garantir que o processo terminou
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()
            return 1
        time.sleep(0.1) # Pequena pausa para não sobrecarregar o CPU

    # Captura qualquer saída restante após o término do processo
    stdout_restante, stderr_output = proc.communicate()
    if stdout_restante:
        print(stdout_restante, end='')
    
    if proc.returncode != 0:
        print(f"   Erro durante a execução de '{os.path.basename(script)}'. Código: {proc.returncode}")
        if stderr_output:
            print(f"   Saída de erro:\n{stderr_output}")
    else:
        print(f" Concluído OK: '{os.path.basename(script)}' para '{nome_arquivo_com_extensao}'")

    return proc.returncode

def processar_documento_unico(script, nome_arquivo_com_extensao, id_revisao):
    """Função wrapper para executar um script para um único documento."""
    print(f"\n Iniciando processamento de '{os.path.basename(script)}' para o documento '{nome_arquivo_com_extensao}'.")
    resultado = rodar_script((script, nome_arquivo_com_extensao, id_revisao))
    if resultado != 0:
        print(f" Houve uma falha durante o processamento de '{os.path.basename(script)}'.")
        return False
    return True

# === Função Principal ===
def main():
    print("--- GERENCIADOR INICIADO COM SUCESSO ---")
    inicio_execucao = time.time()
    
    # MUDANÇA: Parser de argumentos ajustado
    parser = argparse.ArgumentParser(description="Gerenciador de Revisão Dossel")

    parser.add_argument("nome_arquivo", type=str, help="Nome do arquivo .docx na pasta de entrada.")
    parser.add_argument("--id_revisao", type=int, required=True, help="ID da entrada na fila.")
    parser.add_argument("--email", type=str, required=True, help="E-mail do usuário.")
    parser.add_argument("--modo", choices=['completa', 'simples', 'inconsistencias'], required=True, help="O tipo de revisão a ser executada.")
    args = parser.parse_args()

    id_revisao_alvo = args.id_revisao
    email_alvo = args.email
    nome_arquivo_original = args.nome_arquivo
    
    caminho_arquivo_processar = Path(PASTA_ENTRADA) / nome_arquivo_original
    
    if not caminho_arquivo_processar.exists():
        print(f" ERRO CRÍTICO: Arquivo '{nome_arquivo_original}' não encontrado na pasta de entrada.")
        atualizar_status_progresso(100, id_revisao=id_revisao_alvo)
        QueueEntry.objects.filter(id=id_revisao_alvo).update(status='erro')
        return

    nome_base_doc = Path(nome_arquivo_original).stem
    print(f"\n Documento a processar: {nome_arquivo_original}")
    print(f" Modo de revisão: {args.modo}")
    print(f" E-mail destino: {email_alvo}")

    sucesso_geral = True
    atualizar_status_progresso(5, id_revisao=id_revisao_alvo)
    if foi_cancelado(id_revisao_alvo): return

    # A lógica de despacho permanece a mesma
    script_alvo = None
    if args.modo == 'completa':
        script_alvo = SCRIPT_REVISAO_COMPLETA
    elif args.modo == 'simples':
        script_alvo = SCRIPT_REVISAO_SIMPLES
    elif args.modo == 'inconsistencias':
        script_alvo = SCRIPT_VERIFICADOR_INCONSISTENCIAS
    
    if script_alvo:
        # Passamos o nome do arquivo, não o caminho completo.
        if not processar_documento_unico(script_alvo, nome_arquivo_original, id_revisao_alvo):
            sucesso_geral = False
    else:
        print(f"Erro: Modo de revisão '{args.modo}' não reconhecido.")
        sucesso_geral = False

    tempo_total = time.time() - inicio_execucao

    if sucesso_geral:
        QueueEntry.objects.filter(id=id_revisao_alvo).update(duration_seconds=tempo_total)


        atualizar_status_progresso(95, id_revisao=id_revisao_alvo)
        print(" Processamento finalizado com sucesso. Gerando relatórios finais...")
        registrar_documento(nome_base_doc)

        qtd_paginas = obter_contagem_paginas(caminho_arquivo_processar)
        tokens_in, tokens_out, custo_brl = calcular_custo_e_tokens(nome_base_doc, args.modo)
        caminho_saida_relativo = str((Path(PASTA_SAIDA) / nome_base_doc).relative_to(PROJECT_ROOT))

        dados_para_planilha = [
            time.strftime('%Y-%m-%d %H:%M:%S'),
            nome_arquivo_original,
            qtd_paginas,
            args.modo,
            round(tempo_total, 2),
            round(custo_brl, 2),
            caminho_saida_relativo
        ]
        if atualizar_planilha_mestra(dados_para_planilha):
            enviar_planilha_mestra_por_email()
        
        # Lógica para encontrar o arquivo de saída e salvar no BD
        output_file_map = {
            'completa': f"{nome_base_doc}_revisao_completa.docx",
            'simples': f"{nome_base_doc}_revisao_simples.docx",
            'inconsistencias': f"relatorio_tecnico_inconsistencias_{nome_base_doc}.docx"
        }
        output_filename = output_file_map.get(args.modo)
        if output_filename:
            output_path = Path(PASTA_SAIDA) / nome_base_doc / output_filename
            if output_path.exists():
                QueueEntry.objects.filter(id=id_revisao_alvo).update(output_file_url=str(output_path.relative_to(PROJECT_ROOT)))
            else:
                print(f" Arquivo de saída principal '{output_filename}' não encontrado.")

        # Lógica para e-mail do usuário
        if args.modo in ['completa', 'simples']:
            corpo_email = (
                f"Sua revisão '{args.modo}' para o documento '{nome_arquivo_original}' foi concluída com sucesso.\n\n"
                f"Tempo Total de Execução: {tempo_total:.2f} segundos\n"
                f"Custo Total Estimado: R$ {custo_brl:.2f}\n"
            )
        else: # modo 'inconsistencias'
            corpo_email = f"A análise de inconsistências para o documento '{nome_arquivo_original}' foi finalizada em {tempo_total:.2f} segundos. Os relatórios estão disponíveis na sua área de resultados."
        
        enviar_email_final(f"Revisão '{args.modo.upper()}' Concluída: {nome_arquivo_original}", corpo_email, email_alvo)
    else:
        print(" Falhas durante o processamento. O processo pode ter sido cancelado ou encontrado um erro.")

    atualizar_status_progresso(100, id_revisao=id_revisao_alvo)
    print(" Gerenciador finalizou a execução.")
    
if __name__ == "__main__":
    main()